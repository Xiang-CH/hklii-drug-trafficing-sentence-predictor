from __future__ import annotations

import math
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_FILENAME = "sentencing_factor_outlier_lists.xlsx"
INFERRED_ROLE_SOURCE = "Inferred as starting point since role adjustment not provided"
EXCEL_ILLEGAL_CHARACTERS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
FORMULA_PREFIXES = ("=", "+", "-", "@")
DISCOUNT_COMPARISON_DECIMALS = 1
HEADER_BORDER = Border(
	left=Side(style="thin", color="9ECAE1"),
	right=Side(style="thin", color="9ECAE1"),
	top=Side(style="thin", color="9ECAE1"),
	bottom=Side(style="thin", color="9ECAE1"),
)
DATA_BORDER = Border(
	left=Side(style="thin", color="D9E2F3"),
	right=Side(style="thin", color="D9E2F3"),
	top=Side(style="thin", color="D9E2F3"),
	bottom=Side(style="thin", color="D9E2F3"),
)


@dataclass(frozen=True)
class DiscountRule:
	type_name: str
	minimum: float | None
	maximum: float | None
	minimum_inclusive: bool = True
	maximum_inclusive: bool = True

	@property
	def display(self) -> str:
		if self.minimum is not None and self.maximum is not None:
			return f"{self.minimum * 100:.1f}%–{self.maximum * 100:.1f}%"
		if self.maximum is not None:
			comparator = "up to" if self.maximum_inclusive else "below"
			return f"{comparator} {self.maximum * 100:.2f}%"
		return f"at least {self.minimum * 100:.1f}%"


PLEA_RULES = {
	("High Court", "Up to committal"): DiscountRule("Earliest opportunity", 0.313, 0.353),
	("District Court", "Plea day"): DiscountRule("Earliest opportunity", 0.313, 0.353),
	("High Court", "After committal"): DiscountRule("Before trial dates are set", 0.23, 0.27),
	("High Court", "After dates fixed"): DiscountRule("Before trial starts", 0.20, 0.25),
	("District Court", "After dates fixed"): DiscountRule("Before trial starts", 0.20, 0.25),
	("High Court", "First day"): DiscountRule("First day of trial", 0.18, 0.22),
	("District Court", "First day"): DiscountRule("First day of trial", 0.18, 0.22),
	("High Court", "During trial"): DiscountRule("During trial", None, 0.20, maximum_inclusive=False),
	("District Court", "During trial"): DiscountRule("During trial", None, 0.20, maximum_inclusive=False),
}

ASSISTANCE_RULES = {
	"Assistance - limited": DiscountRule("Assistance - limited", None, 0.40),
	"Assistance - useful": DiscountRule("Assistance - useful", 0.40, 0.45),
	"Assistance - testify": DiscountRule("Assistance - testify", 0.48, 0.52),
	"Assistance - risk": DiscountRule("Assistance - risk", None, 2 / 3),
}

STANDARD_SHEETS = {
	"THC-CBD": {
		"description": "Verified charges involving THC/CBD.",
		"match": "thc_cbd",
	},
	"Other-Midazolam": {
		"description": "Verified charges with Other: Midazolam drugs.",
		"match": "midazolam",
	},
	"Import-Export": {
		"description": "Verified charges with an Import or Export aggravating factor.",
		"match": "import_export",
	},
	"Use of minors": {
		"description": "Verified charges with the Use of minors aggravating factor.",
		"match": "use_of_minors",
	},
	"Young offender": {
		"description": "Verified charges with the Young offender mitigating factor.",
		"match": "Young offender",
	},
	"Medical conditions": {
		"description": "Verified charges with the Medical conditions mitigating factor.",
		"match": "Medical conditions",
	},
	"Family illness": {
		"description": "Verified charges with the Family illness mitigating factor.",
		"match": "Family illness",
	},
}

STANDARD_COLUMNS = [
	"neutral_citation",
	"exclude_case",
	"trial_index",
	"Charge_no",
	"Defendant_id",
	"charge_name",
	"drugs",
	"matched_category",
	"matched_values",
	"starting_point_months",
	"starting_point_inferred",
	"sentence_after_role_months",
	"sentence_after_role_inferred",
	"notional_sentence_months",
	"notional_sentence_inferred",
	"mitigation_reduction_months",
	"mitigation_reduction_inferred",
	"final_sentence_months",
	"final_sentence_inferred",
	"remarks",
	"source_text",
]

PLEA_COLUMNS = STANDARD_COLUMNS + [
	"plea_type",
	"court_type",
	"stored_plea_stage",
	"guilty_plea_inferred",
	"actual_plea_discount_pct",
	"expected_discount_rule",
	"outlier_reason",
	"plea_source_text",
]

ASSISTANCE_COLUMNS = STANDARD_COLUMNS + [
	"assistance_type",
	"assistance_inferred",
	"guilty_plea_inferred",
	"assistance_discount_pct",
	"guilty_plea_discount_pct",
	"combined_discount_pct",
	"expected_discount_rule",
	"outlier_reason",
	"other_mitigating_factors",
	"assistance_source_text",
	"plea_source_text",
]


def repo_root() -> Path:
	"""Return the repository root from either notebook or repository CWD."""
	current = Path.cwd().resolve()
	if (current / "featureExtraction").exists():
		return current
	if (current.parent / "featureExtraction").exists():
		return current.parent
	return Path(__file__).resolve().parents[1]


def load_environment(root: Path) -> None:
	for env_path in (
		root / "featureExtraction" / ".env",
		root / "featureVerification" / ".env.local",
		root / ".env",
	):
		if env_path.exists():
			load_dotenv(env_path)


def get_verified_collection():
	root = repo_root()
	load_environment(root)
	notebooks_path = str(root / "notebooks")
	if notebooks_path not in sys.path:
		sys.path.insert(0, notebooks_path)
	from evaluate_verified_sentences import get_collection

	collection, _ = get_collection()
	return collection


def clean_excel_text(value: Any) -> str | None:
	if value is None:
		return None
	text = EXCEL_ILLEGAL_CHARACTERS.sub("", str(value)).strip()
	if not text:
		return None
	return f"'{text}" if text.startswith(FORMULA_PREFIXES) else text


def unique_texts(values: Iterable[Any]) -> str | None:
	seen: set[str] = set()
	result: list[str] = []
	for value in values:
		text = clean_excel_text(value)
		if text and text not in seen:
			seen.add(text)
			result.append(text)
	return " | ".join(result) if result else None


def is_inferred(detail: dict[str, Any] | None) -> bool:
	if not detail:
		return False
	return bool(detail.get("inferred")) or detail.get("source") == INFERRED_ROLE_SOURCE


def total_months(detail: dict[str, Any] | None) -> int | None:
	if not detail:
		return None
	if detail.get("total_months") is not None:
		return int(detail["total_months"])
	if detail.get("sentence_years") is not None or detail.get("sentence_months") is not None:
		return int(detail.get("sentence_years") or 0) * 12 + int(detail.get("sentence_months") or 0)
	return None


def reduction_months(detail: dict[str, Any] | None) -> int | None:
	if not detail:
		return None
	if detail.get("reduction_months") is None:
		return None
	return int(detail["reduction_months"])


def plea_reduction_months(detail: dict[str, Any] | None) -> int | None:
	if not detail:
		return None
	years = detail.get("reduction_years")
	months = detail.get("reduction_months")
	if years is None and months is None:
		return None
	return int(years or 0) * 12 + int(months or 0)


def explicit_discount_fraction(detail: dict[str, Any] | None, base_months: int | None, months_getter=reduction_months) -> float | None:
	if not detail:
		return None
	percentage = detail.get("reduction_percentage")
	if percentage is not None:
		return float(percentage) / 100
	months = months_getter(detail)
	if months is None or base_months is None or base_months <= 0:
		return None
	return months / base_months


def plea_stage(plea: dict[str, Any]) -> str | None:
	return plea.get("high_court_stage") or plea.get("district_court_stage")


def classify_discount(value: float, rule: DiscountRule) -> tuple[bool, str | None]:
	if not math.isfinite(value):
		return True, "Discount is not a finite number."
	# The workbook displays discounts to one decimal percentage point. Compare at
	# that same precision so a displayed boundary value is never flagged solely
	# because of floating-point or source-value precision beyond what reviewers see.
	comparison_value = round(value * 100, DISCOUNT_COMPARISON_DECIMALS) / 100
	if rule.minimum is not None:
		too_low = comparison_value < rule.minimum if rule.minimum_inclusive else comparison_value <= rule.minimum
		if too_low:
			return True, f"{comparison_value:.1%} is below the expected {rule.display}."
	if rule.maximum is not None:
		too_high = comparison_value > rule.maximum if rule.maximum_inclusive else comparison_value >= rule.maximum
		if too_high:
			return True, f"{comparison_value:.1%} is above the expected {rule.display}."
	return False, None


def combined_discount(assistance_fraction: float, plea_fraction: float) -> float:
	return 1 - (1 - assistance_fraction) * (1 - plea_fraction)


def format_drugs(trial: dict[str, Any]) -> str | None:
	parts: list[str] = []
	for drug in trial.get("drugs") or []:
		drug_type = drug.get("drug_type")
		if not drug_type:
			continue
		label = drug_type
		if drug_type == "Other" and drug.get("other_drug_type"):
			label = f"Other: {drug['other_drug_type']}"
		quantity = drug.get("quantity")
		parts.append(f"{label}: {quantity}g" if quantity is not None else label)
	return unique_texts(parts)


def base_row(document: dict[str, Any], trial: dict[str, Any], trial_index: int) -> dict[str, Any]:
	charge = trial.get("charge_type") or {}
	starting_point = trial.get("starting_point")
	sentence_after_role = trial.get("sentence_after_role")
	notional_sentence = trial.get("notional_sentence")
	mitigation_reduction = trial.get("mitigation_reduction")
	final_sentence = trial.get("final_sentence")
	return {
		"neutral_citation": clean_excel_text((document.get("judgement") or {}).get("neutral_citation")),
		"exclude_case": bool(document.get("exclude")),
		"trial_index": trial_index + 1,
		"Charge_no": charge.get("charge_no"),
		"Defendant_id": charge.get("defendant_id"),
		"charge_name": clean_excel_text(charge.get("charge_name")),
		"drugs": format_drugs(trial),
		"matched_category": None,
		"matched_values": None,
		"starting_point_months": total_months(starting_point),
		"starting_point_inferred": is_inferred(starting_point),
		"sentence_after_role_months": total_months(sentence_after_role),
		"sentence_after_role_inferred": is_inferred(sentence_after_role),
		"notional_sentence_months": total_months(notional_sentence),
		"notional_sentence_inferred": is_inferred(notional_sentence),
		"mitigation_reduction_months": reduction_months(mitigation_reduction),
		"mitigation_reduction_inferred": is_inferred(mitigation_reduction),
		"final_sentence_months": total_months(final_sentence),
		"final_sentence_inferred": is_inferred(final_sentence),
		"remarks": clean_excel_text(document.get("remarks")),
		"source_text": None,
	}


def matching_standard_values(trial: dict[str, Any], match: str) -> tuple[str, list[str], list[str]]:
	drugs = trial.get("drugs") or []
	aggravating = trial.get("aggravating_factors") or []
	mitigating = trial.get("mitigating_factors") or []
	if match == "thc_cbd":
		items = [drug for drug in drugs if drug.get("drug_type") == "THC/CBD"]
		return "Drug", ["THC/CBD"] * len(items), [item.get("source") for item in items]
	if match == "midazolam":
		items = [
			drug for drug in drugs
			if drug.get("drug_type") == "Other"
			and str(drug.get("other_drug_type") or "").strip().casefold() == "midazolam"
		]
		return "Drug", ["Other: Midazolam"] * len(items), [item.get("source") for item in items]
	if match == "import_export":
		items = [factor for factor in aggravating if factor.get("factor") in {"Import", "Export"}]
		return "Aggravating factor", [item.get("factor") for item in items], [item.get("source") for item in items]
	if match == "use_of_minors":
		items = [factor for factor in aggravating if factor.get("factor") == "Use of minors"]
		return "Aggravating factor", ["Use of minors"] * len(items), [item.get("source") for item in items]
	items = [factor for factor in mitigating if factor.get("factor") == match]
	return "Mitigating factor", [match] * len(items), [item.get("source") for item in items]


def make_standard_rows(documents: Iterable[dict[str, Any]], match: str) -> list[dict[str, Any]]:
	rows: list[dict[str, Any]] = []
	for document in documents:
		for index, trial in enumerate((document.get("trials") or {}).get("trials") or []):
			category, values, sources = matching_standard_values(trial, match)
			if not values:
				continue
			row = base_row(document, trial, index)
			row.update({
				"matched_category": category,
				"matched_values": unique_texts(values),
				"source_text": "\n\n".join(value for value in (clean_excel_text(source) for source in sources) if value),
			})
			rows.append(row)
	return rows


def plea_discount(trial: dict[str, Any]) -> float | None:
	plea = trial.get("guilty_plea") or {}
	if not plea.get("pleaded_guilty"):
		return None
	pre_plea_base: int | None = None
	notional = total_months(trial.get("notional_sentence"))
	mitigation = reduction_months(trial.get("mitigation_reduction"))
	if notional is not None and mitigation is not None:
		pre_plea_base = notional - mitigation
	return explicit_discount_fraction(plea, pre_plea_base, plea_reduction_months)


def make_plea_outlier_rows(documents: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
	rows: list[dict[str, Any]] = []
	not_assessable = 0
	for document in documents:
		for index, trial in enumerate((document.get("trials") or {}).get("trials") or []):
			plea = trial.get("guilty_plea") or {}
			if not plea.get("pleaded_guilty"):
				continue
			court = plea.get("court_type")
			stage = plea_stage(plea)
			rule = PLEA_RULES.get((court, stage))
			if rule is None:
				continue
			discount = plea_discount(trial)
			if discount is None:
				not_assessable += 1
				continue
			is_outlier, reason = classify_discount(discount, rule)
			if not is_outlier:
				continue
			row = base_row(document, trial, index)
			row.update({
				"matched_category": "Guilty plea",
				"matched_values": rule.type_name,
				"source_text": clean_excel_text(plea.get("source")),
				"plea_type": rule.type_name,
				"court_type": clean_excel_text(court),
				"stored_plea_stage": clean_excel_text(stage),
				"guilty_plea_inferred": is_inferred(plea),
				"actual_plea_discount_pct": discount,
				"expected_discount_rule": rule.display,
				"outlier_reason": reason,
				"plea_source_text": clean_excel_text(plea.get("source")),
			})
			rows.append(row)
	return rows, not_assessable


def make_assistance_outlier_rows(documents: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
	rows: list[dict[str, Any]] = []
	not_assessable = 0
	for document in documents:
		for index, trial in enumerate((document.get("trials") or {}).get("trials") or []):
			notional = total_months(trial.get("notional_sentence"))
			plea = trial.get("guilty_plea") or {}
			plea_fraction = plea_discount(trial)
			mitigating = trial.get("mitigating_factors") or []
			for factor in mitigating:
				factor_name = factor.get("factor")
				rule = ASSISTANCE_RULES.get(factor_name)
				if rule is None:
					continue
				assistance_fraction = explicit_discount_fraction(factor, notional)
				if assistance_fraction is None or plea_fraction is None:
					not_assessable += 1
					continue
				combined = combined_discount(assistance_fraction, plea_fraction)
				is_outlier, reason = classify_discount(combined, rule)
				if not is_outlier:
					continue
				other_factors = [
					item.get("factor") for item in mitigating
					if item.get("factor") and item is not factor
				]
				row = base_row(document, trial, index)
				row.update({
					"matched_category": "Mitigating factor",
					"matched_values": factor_name,
					"source_text": unique_texts([factor.get("source"), plea.get("source")]),
					"assistance_type": factor_name,
					"assistance_inferred": is_inferred(factor),
					"guilty_plea_inferred": is_inferred(plea),
					"assistance_discount_pct": assistance_fraction,
					"guilty_plea_discount_pct": plea_fraction,
					"combined_discount_pct": combined,
					"expected_discount_rule": rule.display,
					"outlier_reason": reason,
					"other_mitigating_factors": unique_texts(other_factors),
					"assistance_source_text": clean_excel_text(factor.get("source")),
					"plea_source_text": clean_excel_text(plea.get("source")),
				})
				rows.append(row)
	return rows, not_assessable


def frame(rows: list[dict[str, Any]], columns: list[str]) -> pd.DataFrame:
	return pd.DataFrame(rows, columns=columns).sort_values(
		["neutral_citation", "trial_index", "Charge_no", "Defendant_id"],
		na_position="last",
	).reset_index(drop=True)


def assert_synthetic_rules() -> None:
	assert matching_standard_values({"drugs": [{"drug_type": "THC/CBD"}]}, "thc_cbd")[1] == ["THC/CBD"]
	assert matching_standard_values({"drugs": [{"drug_type": "Other", "other_drug_type": "  MiDaZoLaM "}]}, "midazolam")[1] == ["Other: Midazolam"]
	assert PLEA_RULES[("High Court", "After committal")].type_name == "Before trial dates are set"
	assert not classify_discount(0.313, PLEA_RULES[("High Court", "Up to committal")])[0]
	assert classify_discount(0.3124, PLEA_RULES[("High Court", "Up to committal")])[0]
	assert not classify_discount(0.20, PLEA_RULES[("High Court", "After dates fixed")])[0]
	assert not classify_discount(0.25, PLEA_RULES[("High Court", "After dates fixed")])[0]
	assert classify_discount(0.199, PLEA_RULES[("High Court", "After dates fixed")])[0]
	assert not classify_discount(0.1994, PLEA_RULES[("High Court", "During trial")])[0]
	assert classify_discount(0.20, PLEA_RULES[("High Court", "During trial")])[0]
	assert math.isclose(combined_discount(0.10, 1 / 3), 0.40)
	assert not classify_discount(0.40, ASSISTANCE_RULES["Assistance - limited"])[0]
	assert classify_discount(0.4006, ASSISTANCE_RULES["Assistance - limited"])[0]
	assert not classify_discount(0.48, ASSISTANCE_RULES["Assistance - testify"])[0]
	assert classify_discount(0.521, ASSISTANCE_RULES["Assistance - testify"])[0]
	assert not classify_discount(0.3533, PLEA_RULES[("High Court", "Up to committal")])[0]
	assert not classify_discount(0.479974, ASSISTANCE_RULES["Assistance - testify"])[0]
	assert total_months({"sentence_years": 1, "inferred": True}) == 12
	assert math.isclose(explicit_discount_fraction({"reduction_percentage": 25, "inferred": True}, None), 0.25)


def excel_value(value: Any) -> Any:
	if value is None or (isinstance(value, float) and math.isnan(value)):
		return None
	if isinstance(value, (str, int, float, bool)):
		return value
	return clean_excel_text(value)


def write_dataframe_sheet(workbook: Workbook, name: str, description: str, dataframe: pd.DataFrame) -> None:
	worksheet = workbook.create_sheet(name)
	worksheet.sheet_view.showGridLines = False
	worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(dataframe.columns)))
	worksheet.cell(1, 1, name)
	worksheet.cell(1, 1).font = Font(name="Aptos Display", size=15, bold=True, color="FFFFFF")
	worksheet.cell(1, 1).fill = PatternFill("solid", fgColor="1F4E78")
	worksheet.cell(1, 1).alignment = Alignment(vertical="center")
	worksheet.row_dimensions[1].height = 26
	worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(dataframe.columns)))
	worksheet.cell(2, 1, description)
	worksheet.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
	worksheet.row_dimensions[2].height = 32
	header_row = 3
	for column_index, column in enumerate(dataframe.columns, start=1):
		cell = worksheet.cell(header_row, column_index, column)
		cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
		cell.fill = PatternFill("solid", fgColor="2F75B5")
		cell.border = HEADER_BORDER
		cell.alignment = Alignment(wrap_text=True, vertical="center")
	for row in dataframe.itertuples(index=False, name=None):
		worksheet.append([excel_value(value) for value in row])
	last_row = header_row + len(dataframe)
	last_column = max(1, len(dataframe.columns))
	worksheet.freeze_panes = "A4"
	if len(dataframe.columns):
		worksheet.auto_filter.ref = f"A{header_row}:{worksheet.cell(last_row, last_column).coordinate}"
	for column_index, column in enumerate(dataframe.columns, start=1):
		letter = get_column_letter(column_index)
		width = min(max(len(column) + 2, 12), 28)
		if column in {"source_text", "plea_source_text", "assistance_source_text", "remarks", "outlier_reason"}:
			width = 55
		elif column in {"drugs", "other_mitigating_factors", "expected_discount_rule"}:
			width = 30
		worksheet.column_dimensions[letter].width = width
		for cell in worksheet[letter][header_row:]:
			if cell.row > header_row:
				cell.border = DATA_BORDER
			cell.alignment = Alignment(vertical="top", wrap_text=column in {"source_text", "plea_source_text", "assistance_source_text", "remarks", "outlier_reason", "drugs", "other_mitigating_factors"})
			if column.endswith("_pct"):
				cell.number_format = "0.0%"
		for cell in worksheet[letter][header_row + 1:]:
			if cell.alignment.wrap_text:
				cell.alignment = Alignment(vertical="top", wrap_text=True)
	for row_index in range(header_row + 1, last_row + 1):
		worksheet.row_dimensions[row_index].height = 32


def write_summary_sheet(workbook: Workbook, summary_rows: list[tuple[str, int, str]]) -> None:
	dataframe = pd.DataFrame(summary_rows, columns=["metric", "count", "notes"])
	write_dataframe_sheet(
		workbook,
		"Summary",
		"Generated " + datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z") + ". Inferred sentence data is included and identified by *_inferred columns. Records without enough quantitative data are counted but not listed as outliers.",
		dataframe,
	)
	worksheet = workbook["Summary"]
	worksheet.column_dimensions["A"].width = 34
	worksheet.column_dimensions["B"].width = 16
	worksheet.column_dimensions["C"].width = 82
	for row_index in range(5, worksheet.max_row + 1):
		worksheet.cell(row_index, 3).alignment = Alignment(vertical="top", wrap_text=True)
		worksheet.row_dimensions[row_index].height = 42


def export_workbook(output_path: Path, sheets: dict[str, pd.DataFrame], plea_not_assessable: int, assistance_not_assessable: int, document_count: int, charge_count: int) -> None:
	workbook = Workbook()
	workbook.remove(workbook.active)
	summary_rows = [
		("Verified judgments processed", document_count, "Query filter: is_verified=True; excluded cases remain visible in outputs."),
		("Verified charges processed", charge_count, "One row per trial/charge-defendant record."),
	]
	for name, dataframe in sheets.items():
		summary_rows.append((name, len(dataframe), "Rows written to this worksheet."))
	summary_rows.extend([
		("Plea records not assessable", plea_not_assessable, "Mapped plea stage but missing quantitative reduction data."),
		("Assistance records not assessable", assistance_not_assessable, "Assistance or guilty-plea component has missing quantitative data."),
	])
	write_summary_sheet(workbook, summary_rows)
	for name, dataframe in sheets.items():
		if name in STANDARD_SHEETS:
			description = STANDARD_SHEETS[name]["description"]
		elif name == "Plea outliers":
			description = "Outliers across all five specified guilty-plea timings. Use plea_type to classify the timing."
		else:
			description = "Outliers across all four assistance types. Use assistance_type to classify the assistance category."
		write_dataframe_sheet(workbook, name, description, dataframe)
	output_path.parent.mkdir(parents=True, exist_ok=True)
	workbook.save(output_path)


def verify_workbook(output_path: Path, sheets: dict[str, pd.DataFrame]) -> None:
	workbook = load_workbook(output_path, read_only=False, data_only=False)
	expected_names = ["Summary", *sheets]
	assert workbook.sheetnames == expected_names, f"Unexpected worksheets: {workbook.sheetnames}"
	for name, dataframe in sheets.items():
		worksheet = workbook[name]
		headers = [worksheet.cell(3, index).value for index in range(1, len(dataframe.columns) + 1)]
		assert headers == list(dataframe.columns), f"Unexpected headers in {name}"
		assert worksheet.max_row == 3 + len(dataframe), f"Unexpected row count in {name}"
		expected_filter = f"A3:{worksheet.cell(3 + len(dataframe), len(dataframe.columns)).coordinate}"
		assert worksheet.auto_filter.ref == expected_filter, f"Unexpected filter range in {name}"
		assert not worksheet.tables, f"Unexpected tables in {name}"


def build_workbook(documents: list[dict[str, Any]], output_path: Path) -> dict[str, int]:
	assert_synthetic_rules()
	standard_frames = {
		name: frame(make_standard_rows(documents, config["match"]), STANDARD_COLUMNS)
		for name, config in STANDARD_SHEETS.items()
	}
	plea_rows, plea_not_assessable = make_plea_outlier_rows(documents)
	assistance_rows, assistance_not_assessable = make_assistance_outlier_rows(documents)
	sheets = {
		"THC-CBD": standard_frames["THC-CBD"],
		"Other-Midazolam": standard_frames["Other-Midazolam"],
		"Import-Export": standard_frames["Import-Export"],
		"Plea outliers": frame(plea_rows, PLEA_COLUMNS),
		"Use of minors": standard_frames["Use of minors"],
		"Assistance outliers": frame(assistance_rows, ASSISTANCE_COLUMNS),
		"Young offender": standard_frames["Young offender"],
		"Medical conditions": standard_frames["Medical conditions"],
		"Family illness": standard_frames["Family illness"],
	}
	charge_count = sum(len((document.get("trials") or {}).get("trials") or []) for document in documents)
	export_workbook(output_path, sheets, plea_not_assessable, assistance_not_assessable, len(documents), charge_count)
	verify_workbook(output_path, sheets)
	return {
		"verified_judgments": len(documents),
		"verified_charges": charge_count,
		"plea_outliers": len(sheets["Plea outliers"]),
		"plea_not_assessable": plea_not_assessable,
		"assistance_outliers": len(sheets["Assistance outliers"]),
		"assistance_not_assessable": assistance_not_assessable,
		**{f"rows_{name}": len(dataframe) for name, dataframe in sheets.items()},
	}


def main() -> dict[str, int]:
	collection = get_verified_collection()
	projection = {
		"exclude": 1,
		"remarks": 1,
		"judgement.neutral_citation": 1,
		"trials": 1,
	}
	documents = list(collection.find({"is_verified": True}, projection))
	output_path = repo_root() / "notebooks" / OUTPUT_FILENAME
	result = build_workbook(documents, output_path)
	print(f"Wrote {output_path}")
	for key, value in result.items():
		print(f"{key.replace('_', ' ')}: {value}")
	return result


if __name__ == "__main__":
	main()
